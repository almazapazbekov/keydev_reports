from typing import Iterable, NoReturn

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseEditor


class ExcelEditor(BaseEditor):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        """
        Конструктор класса. Принимает имя файла и загружает его.
        :param file_name: Имя Excel-файла.
        """
        self.workbook = load_workbook(self.file_name)

    @staticmethod
    def find_target_cell(sheet: Worksheet, search_text: str) -> Cell:
        """
        Статический метод для поиска ячейки с заданным текстом на листе.
        :param sheet: Лист, на котором ищется ячейка.
        :param search_text: Текст для поиска.
        :return: Найденная ячейка.
        """
        target_cell = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == search_text and cell.value is not None:
                    target_cell = cell
                    break
        return target_cell

    @staticmethod
    def fill_data(sheet: Worksheet, data: list, start_row: int, start_col: int) -> NoReturn:
        """
        Статический метод для заполнения данных на листе, начиная с указанной ячейки.
        :param sheet: Лист, на котором заполняются данные.
        :param data: Данные для вставки.
        :param start_row: Начальная строка.
        :param start_col: Начальная колонка.
        """
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, value in enumerate(row_data, start=start_col):
                cell = sheet.cell(row=row_index, column=col_index)
                if cell.data_type == 'f':
                    continue
                else:
                    cell.value = value

    def replace_in_excel(self, search_text: str, replace_text: str) -> NoReturn:
        """
        Метод для поиска и замены текста в ячейках Excel-файла.
        :param search_text: Текст для поиска.
        :param replace_text: Текст для замены.
        :return: Excel-файл с выполненными заменами.
        """
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell.value, Iterable) and search_text in cell.value:
                        cell.value = cell.value.replace(search_text, replace_text)


    @staticmethod
    def shift_cells(sheet: Worksheet,  start_row: int, shift_length: int) -> NoReturn:
        """
        Начинает работу с заданной строки (start_row)
        Двигает вниз объединенные строки со всеми данными в нем и создаёт пустые строки,
        это сделано для того, чтобы не ломалась структура документа,

        :param sheet: Лист.
        :param start_row: Получает текущее положение строки, начиная с которой идет переформатирование.
        :param shift_length: Длинна объекта данных, полученных в data (Кол-во строк, на которое он сдвинет все вниз).

        """
        merged_ranges = sheet.merged_cells.ranges
        for merged_cell in merged_ranges:
            _, min_row, _, _ = range_boundaries(merged_cell.coord)
            if min_row >= start_row:
                merged_cell.shift(0, shift_length)

        sheet.insert_rows(start_row, shift_length)

    def save_excel(self, new_file_name=None) -> NoReturn:
        """
        Метод для сохранения Excel-файла с новым именем (если указано) или перезаписи текущего.
        :param new_file_name: Новое имя файла (по умолчанию перезапись текущего файла).
        """
        self.file_name = new_file_name if new_file_name else self.file_name
        self.workbook.save(self.file_name)
